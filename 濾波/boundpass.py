# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

from scipy import signal
import xlrd
import os
from openpyxl import load_workbook
import numpy as np


data=[]
# 讀取 Excel 檔案
wb = load_workbook('HW2_2.xlsx')
sheet = wb['工作表1']
for i in range(32768):
    cell = sheet.cell(row=i+1, column=1).value
    data.append(cell)
    print(i)
    
x= np.array(data) 

b, a = signal.butter(8, [0.001,0.3], 'bandpass')   #配置濾波器 8 表示濾波器的階數
filtedData = signal.filtfilt(b, a, x)  #data為要過濾的訊號